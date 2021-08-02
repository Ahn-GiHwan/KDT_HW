# 메일을 발송 기능을 위한 모듈 import
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# NaverNewsCrawler 모듈 import
from NaverNewsCrawler import NaverNewsCrawler

# excel파일 읽기 기능을 위한 모듈 import
from openpyxl import load_workbook

# 사용자로 부터 기사 수집을 위한 keyword 입력받기
keyword = input("키워드 : ")
# NaverNewsCrawler class 생성 + 생성자에 값(keyword) 대입
crawler = NaverNewsCrawler(keyword)

# 사용자로 부터 저장할 파일의 이름을 위한 name 입력받기
fileName = input("파일 이름 : ")
# 받은 입력을 NaverNewsCrawler로 전달 => 받은 입력으로 파일 생성
crawler.get_news(fileName)

# SMTP, 이메일 전송을 위한 설정
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USER = 'KDTfrontend@gmail.com'
SMTP_PASSWORD = '@KDT123456'

# 이메일 전송을 위한 함수


def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition',
                             'attachment; filename="'+filename+'"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()


# load_workbook class 생성 + 읽은 excel의 주소를 생성자 매개변수로 전달
wb = load_workbook('./email_list.xlsx')

# 활성화된 시트 할당
data = wb.active

# 읽을 cell 범위를 설정
area = data['B3:C4']

# 이메일 제목
title = "자동화 메일입니다."

# 내용에 들어갈 문구
contents = "[" + keyword + "]에 대한 기사들입니다."

# 이름과 이메일 주소만 가져와서 send_mail함수의 매개변수로 넣어준다.
count = 0
for row in area:
    for cell in row:
        count += 1
        if(count/2 == 1):
            send_mail(name, cell.value, title, contents, fileName + '.xlsx')
            count = 0
        name = cell.value
